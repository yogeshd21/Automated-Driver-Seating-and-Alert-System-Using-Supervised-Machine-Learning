import cv2
import numpy as np
from PIL import Image
import os
import pandas
from openpyxl import load_workbook
import xlsxwriter

from scipy.spatial import distance
from imutils import face_utils
import imutils
import dlib
#import cv2
import playsound

import cairo
import math
import matplotlib.pyplot as plt
import matplotlib.image as mpimg

modeclear = int(input('Press 1 to clear data and have new setup '))
moderec = int(input('Press 1 for new user setup or update '))
#for complete new setup
if modeclear == 1:
    #print('1')
    answerbook = xlsxwriter.Workbook('output.xlsx')
    ans = answerbook.add_worksheet()
    ans.write(0, 0, 'id')
    ans.write(0, 1, 'name')
    ans.write(0, 2, 'seat_type')
    answerbook.close()

#names = ['None']
# data generation
if moderec == 1:
    #print('1 or 2')
    cam = cv2.VideoCapture(0, cv2.CAP_DSHOW)
    cam.set(3, 640)  # set video width
    cam.set(4, 480)  # set video height

    face_detector = cv2.CascadeClassifier('haarcascade_frontalface_default.xml')

    drivers = int(input('Enter the number of drivers to train'))

    for i in range(drivers):
        # For each person, enter one numeric face id
        face_id = int(input('\n Enter user id '))
        dvrnm = input('\n Enter name of the user ')
        seattyp = input('\n Enter seat type ')

        df = pandas.DataFrame({'id': [face_id], 'name': [dvrnm], 'seat_type': [seattyp]})
        writer = pandas.ExcelWriter('output.xlsx', engine='openpyxl', mode='a')
        writer.book = load_workbook('output.xlsx')
        writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
        df.to_excel(writer, index=False, header=False, startrow=face_id)
        writer.save()
        #names.append(dvrnm)

        print("\n [INFO] Initializing face capture. Look the camera and wait ...")
        # Initialize individual sampling face count
        count = 0

        while (True):

            ret, img = cam.read()
            # img = cv2.flip(img, -1) # flip video image vertically
            gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
            faces = face_detector.detectMultiScale(gray, 1.3, 5)

            for (x, y, w, h) in faces:
                cv2.rectangle(img, (x, y), (x + w, y + h), (255, 0, 0), 2)
                count += 1

                # Save the captured image into the datasets folder
                cv2.imwrite("dataset/User." + str(face_id) + '.' + str(count) + ".jpg", gray[y:y + h, x:x + w])

                cv2.imshow('image', img)

            k = cv2.waitKey(100) & 0xff  # Press 'ESC' for exiting video
            if k == 27:
                break
            elif count >= 50:  # Take 50 face sample and stop video
                break

    # Do a bit of cleanup
    print("\n [INFO] Exiting Program and cleanup stuff")
    cam.release()
    cv2.destroyAllWindows()

    # Path for face image database
    path = 'C:/Users/Yogesh/PycharmProjects/pythonProject/Yotry/dataset'
    #path = "dataset/"

    recognizer = cv2.face.LBPHFaceRecognizer_create()  ## for python3
    # recognizer = cv2.face.createLBPHFaceRecognizer()
    detector = cv2.CascadeClassifier("haarcascade_frontalface_default.xml")


    # function to get the images and label data, this is where our generated data is trained
    def getImagesAndLabels(path):

        imagePaths = [os.path.join(path, f) for f in os.listdir(path)]
        faceSamples = []
        ids = []

        for imagePath in imagePaths:

            PIL_img = Image.open(imagePath).convert('L')  # convert it to grayscale
            img_numpy = np.array(PIL_img, 'uint8')

            id = int(os.path.split(imagePath)[-1].split(".")[1])
            faces = detector.detectMultiScale(img_numpy)

            for (x, y, w, h) in faces:
                faceSamples.append(img_numpy[y:y + h, x:x + w])
                ids.append(id)

        return faceSamples, ids


    print("\n [INFO] Training faces. It will take a few seconds. Wait ...")
    faces, ids = getImagesAndLabels(path)
    recognizer.train(faces, np.array(ids))

    # Save the model into trainer/trainer.yml
    recognizer.save('trainer/trainer.yml')  # recognizer.save() worked on Mac, but not on Pi
    # recognizer.save(os.path.join(path, ))

    # Print the numer of faces trained and end program
    print("\n [INFO] {0} faces trained. Exiting Program".format(len(np.unique(ids))))

# Recognition Begins here

recognizer = cv2.face.LBPHFaceRecognizer_create() ## for python3
#recognizer = cv2.face.createLBPHFaceRecognizer()

recognizer.read('trainer/trainer.yml') ## python3
#recognizer.load('trainer/trainer.yml')
cascadePath = "haarcascade_frontalface_default.xml"
faceCascade = cv2.CascadeClassifier(cascadePath)

font = cv2.FONT_HERSHEY_SIMPLEX

# iniciate id counter
id = 0

# names related to ids: example ==> Marcelo: id=1,  etc
#names = ['None', 'YoSpects', 'YoNoSpects', 'Test']
df = pandas.read_excel('output.xlsx', "Sheet1")
nm = list(df['name'].values)
names = ['None']+nm
print(names)

# Initialize and start realtime video capture
cam = cv2.VideoCapture(0, cv2.CAP_DSHOW)
cam.set(3, 640)  # set video widht
cam.set(4, 480)  # set video height

# Define min window size to be recognized as a face
minW = 0.1 * cam.get(3)
minH = 0.1 * cam.get(4)

count = 0

while True:

    ret, img = cam.read()
    # img = cv2.flip(img, -1) # Flip vertically

    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

    faces = faceCascade.detectMultiScale(
        gray,
        scaleFactor=1.2,
        minNeighbors=5,
        minSize=(int(minW), int(minH)),
    )

    for (x, y, w, h) in faces:

        cv2.rectangle(img, (x, y), (x + w, y + h), (0, 255, 0), 2)

        ids, confidence = recognizer.predict(gray[y:y + h, x:x + w])

        # Check if confidence is less them 100 ==> "0" is perfect match
        if (confidence < 70):
            id = names[ids]
            confidence = "  {0}%".format(round(100 - confidence)) #if 30% or more confident identity is true else unknown
        else:
            id = "unknown"
            confidence = "  {0}%".format(round(100 - confidence))

        cv2.putText(img, str(id), (x + 5, y - 5), font, 1, (255, 255, 255), 2)
        cv2.putText(img, str(confidence), (x + 5, y + h - 5), font, 1, (255, 255, 0), 1)

    cv2.imshow('camera', img)

    k = cv2.waitKey(10) & 0xff  # Press 'ESC' for exiting video
    if k == 27:
        break
    count+=1
    if count == 200:
        break

# Do a bit of cleanup
print("\n [INFO] Exiting Program and cleanup stuff")
cam.release()
cv2.destroyAllWindows()

print(id)
st = [0]+list(df['seat_type'])
for i in range(df.shape[0]+1):
    #print(names[i])
    if str(names[i]) == str(id):
        seat = st[i]
        break
    else:
        seat = 0

# pass seat variable to tkinter
print(seat)
def tk(seat):

    if seat == 0:
        surface = cairo.ImageSurface(cairo.FORMAT_ARGB32, 550, 550)
        ctxt = cairo.Context(surface)
        ctxt.translate(95, 98.5)
        ctxt.scale(1.2, 5)
        ctxt.arc(10.0, 10.0, 20.0, 3 * math.pi / 2, math.pi / 2)
        ctxt.set_source_rgba(0, 0, 0, 1)
        ctxt.fill_preserve()
        ctxt.stroke()

        ctxt.translate(38, 28.5)
        ctxt.scale(4.5, 0.5)
        ctxt.arc(5.0, 5.0, 10.0, math.pi, 0)
        ctxt.set_source_rgba(0, 0, 0, 1)
        ctxt.fill_preserve()
        ctxt.stroke()

        ctxt.translate(25, 0)
        ctxt.set_line_cap(cairo.LINE_CAP_ROUND)
        ctxt.move_to(10, 10)
        ctxt.line_to(3, 35)
        ctxt.stroke()

        ctxt.translate(0, -60)
        ctxt.scale(0.7, 3)
        ctxt.arc(5.0, 5.0, 6.0, math.pi / 2, 3 * math.pi / 2)
        ctxt.set_source_rgba(0, 0, 0, 1)
        ctxt.fill_preserve()
        ctxt.stroke()


        surface.write_to_png("ex1.png")
        img = mpimg.imread("ex1.png")
        plt.imshow(img)
        plt.show()

    elif seat == 1:
        surface = cairo.ImageSurface(cairo.FORMAT_ARGB32, 550, 550)
        ctxt = cairo.Context(surface)
        ctxt.save()
        ctxt.translate(95, 98.5)
        ctxt.rotate(-math.pi / 10)
        ctxt.scale(1.5, 5)
        ctxt.arc(10.0, 10.0, 20.0, 3 * math.pi / 2, math.pi / 2)
        ctxt.set_source_rgba(0, 0, 0, 1)
        ctxt.fill_preserve()
        ctxt.stroke()
        ctxt.restore()

        ctxt.translate(191, 235)
        ctxt.scale(5.5, 1)
        ctxt.arc(5.0, 5.0, 10.0, math.pi, 0)
        ctxt.set_source_rgba(0, 0, 0, 1)
        ctxt.fill_preserve()
        ctxt.stroke()

        ctxt.translate(25, -30)
        ctxt.scale(0.9, 3.5)
        ctxt.set_line_cap(cairo.LINE_CAP_ROUND)
        ctxt.move_to(10, 10)
        ctxt.line_to(3, 30)
        ctxt.stroke()

        ctxt.translate(-5, -45)
        ctxt.scale(0.9, 2.5)
        ctxt.arc(5.0, 5.0, 5.0, math.pi / 2, 3 * math.pi / 2)
        ctxt.set_source_rgba(0, 0, 0, 1)
        ctxt.fill_preserve()
        ctxt.stroke()

        surface.write_to_png("ex_std.png")

    #        surface = cairo.ImageSurface(cairo.FORMAT_ARGB32, 550, 550)
#        ctxt = cairo.Context(surface)
#        ctxt.translate(95, 98.5)
#        ctxt.scale(0.5, 5)
#        ctxt.arc(10.0, 10.0, 20.0, 3 * math.pi / 2, math.pi / 2)
#        ctxt.set_source_rgba(0, 0, 0, 1)
#        ctxt.fill_preserve()
#        ctxt.stroke()
#
#        ctxt.translate(50, 28.5)
#        ctxt.scale(6.5, 0.5)
#        ctxt.arc(5.0, 5.0, 10.0, math.pi, 0)
#        ctxt.set_source_rgba(0, 0, 0, 1)
#        ctxt.fill_preserve()
#        ctxt.stroke()
#
#        ctxt.translate(30, -10)
#        ctxt.set_line_cap(cairo.LINE_CAP_ROUND)
#        ctxt.move_to(10, 10)
#        ctxt.line_to(3, 30)
#        ctxt.stroke()
#
#        ctxt.translate(0, -60)
#        ctxt.scale(0.7, 3)
#        ctxt.arc(5.0, 5.0, 5.0, math.pi / 2, 3 * math.pi / 2)
#        ctxt.set_source_rgba(0, 0, 0, 1)
#        ctxt.fill_preserve()
#        ctxt.stroke()
#
#        surface.write_to_png("ex_std.png")
        img = mpimg.imread("ex_std.png")
        plt.imshow(img)
        plt.show()

    elif seat == 2:
        surface = cairo.ImageSurface(cairo.FORMAT_ARGB32, 550, 550)
        ctxt = cairo.Context(surface)
        ctxt.save()
        ctxt.translate(80, 107)
        ctxt.rotate(-math.pi / 16)
        ctxt.scale(1.5, 5)
        ctxt.arc(10.0, 10.0, 20.0, 3 * math.pi / 2, math.pi / 2)
        ctxt.set_source_rgba(0, 0, 0, 1)
        ctxt.fill_preserve()
        ctxt.stroke()
        ctxt.restore()

        ctxt.translate(160, 246)
        ctxt.scale(12 / 2.0, 5 / 2.0)
        ctxt.arc(5.0, 5.0, 10.0, math.pi, 0)
        ctxt.set_source_rgba(0, 0, 0, 1)
        ctxt.fill_preserve()
        ctxt.stroke()

        ctxt.translate(20, -5)
        ctxt.set_line_cap(cairo.LINE_CAP_ROUND)
        ctxt.move_to(10, 10)
        ctxt.line_to(3, 35)
        ctxt.stroke()

        ctxt.translate(-5, -60)
        ctxt.scale(0.7, 3)
        ctxt.arc(5.0, 5.0, 6.0, math.pi / 2, 3 * math.pi / 2)
        ctxt.set_source_rgba(0, 0, 0, 1)
        ctxt.fill_preserve()
        ctxt.stroke()
        surface.write_to_png("ex2.png")
        img = mpimg.imread("ex2.png")
        plt.imshow(img)
        plt.show()

    elif seat == 3:
        surface = cairo.ImageSurface(cairo.FORMAT_ARGB32, 550, 550)
        ctxt = cairo.Context(surface)
        ctxt.translate(95, 98.5)
        ctxt.scale(1.5, 5)
        ctxt.arc(10.0, 10.0, 20.0, 3 * math.pi / 2, math.pi / 2)
        ctxt.set_source_rgba(0, 0, 0, 1)
        ctxt.fill_preserve()
        ctxt.stroke()

        ctxt.translate(35, 27.5)
        ctxt.scale(4, 0.7)
        ctxt.arc(5.0, 5.0, 10.0, math.pi, 0)
        ctxt.set_source_rgba(0, 0, 0, 1)
        ctxt.fill_preserve()
        ctxt.stroke()

        ctxt.translate(15, -10)
        ctxt.set_line_cap(cairo.LINE_CAP_ROUND)
        ctxt.move_to(20, 10)
        ctxt.line_to(13, 30)
        ctxt.stroke()

        ctxt.translate(7, -40)
        ctxt.scale(0.9, 3)
        ctxt.arc(5.0, 5.0, 4.0, math.pi / 2, 3 * math.pi / 2)
        ctxt.set_source_rgba(0, 0, 0, 1)
        ctxt.fill_preserve()
        ctxt.stroke()
        surface.write_to_png("ex3.png")
        img = mpimg.imread("ex3.png")
        plt.imshow(img)
        plt.show()

    elif seat == 4:
        surface = cairo.ImageSurface(cairo.FORMAT_ARGB32, 550, 550)
        ctxt = cairo.Context(surface)
        ctxt.save()
        ctxt.translate(75, 98.5)
        ctxt.scale(2, 5)
        ctxt.arc(10.0, 10.0, 20.0, 3 * math.pi / 2, math.pi / 2)
        ctxt.set_source_rgba(0, 0, 0, 1)
        ctxt.fill_preserve()
        ctxt.stroke()
        ctxt.restore()

        ctxt.translate(125, 245)
        ctxt.scale(5, 2.2)
        ctxt.arc(5.0, 5.0, 10.0, math.pi, 0)
        ctxt.set_source_rgba(0, 0, 0, 1)
        ctxt.fill_preserve()
        ctxt.stroke()

        ctxt.translate(15, -10)
        ctxt.set_line_cap(cairo.LINE_CAP_ROUND)
        ctxt.move_to(20, 10)
        ctxt.line_to(13, 30)
        ctxt.stroke()

        ctxt.translate(5, -60)
        ctxt.scale(0.9, 3)
        ctxt.arc(5.0, 5.0, 6.0, math.pi / 2, 3 * math.pi / 2)
        ctxt.set_source_rgba(0, 0, 0, 1)
        ctxt.fill_preserve()
        ctxt.stroke()

        surface.write_to_png("ex4.png")
        img = mpimg.imread("ex4.png")
        plt.imshow(img)
        plt.show()

tk(seat)


def eye_aspect_ratio(eye):
    A = distance.euclidean(eye[1], eye[5])
    B = distance.euclidean(eye[2], eye[4])
    C = distance.euclidean(eye[0], eye[3])
    ear = (A + B) / (2.0 * C)
    return ear


thresh = 0.25
frame_check = 20
detect = dlib.get_frontal_face_detector()
predict = dlib.shape_predictor(".\shape_predictor_68_face_landmarks.dat")  # Dat file is the crux of the code

(lStart, lEnd) = face_utils.FACIAL_LANDMARKS_68_IDXS["left_eye"]
(rStart, rEnd) = face_utils.FACIAL_LANDMARKS_68_IDXS["right_eye"]
cap = cv2.VideoCapture(0)
flag = 0
while True:
    ret, frame = cap.read()
    frame = imutils.resize(frame, width=450)
    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
    subjects = detect(gray, 0)
    for subject in subjects:
        shape = predict(gray, subject)
        shape = face_utils.shape_to_np(shape)  # converting to NumPy Array
        leftEye = shape[lStart:lEnd]
        rightEye = shape[rStart:rEnd]
        leftEAR = eye_aspect_ratio(leftEye)
        rightEAR = eye_aspect_ratio(rightEye)
        ear = (leftEAR + rightEAR) / 2.0
        leftEyeHull = cv2.convexHull(leftEye)
        rightEyeHull = cv2.convexHull(rightEye)
        cv2.drawContours(frame, [leftEyeHull], -1, (0, 255, 0), 1)
        cv2.drawContours(frame, [rightEyeHull], -1, (0, 255, 0), 1)
        if ear < thresh:
            flag += 1
            print(flag)
            if flag >= frame_check:
                cv2.putText(frame, "****************ALERT!****************", (10, 30),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 0, 255), 2)
                cv2.putText(frame, "****************ALERT!****************", (10, 325),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 0, 255), 2)
                playsound.playsound('C:/Users/Yogesh/PycharmProjects/pythonProject/Yotry/alarm.wav')
            # print ("Drowsy")
        else:
            flag = 0
    cv2.imshow("Frame", frame)
    key = cv2.waitKey(1) & 0xFF
    if key == ord("q"):
        break
    if key == ord("s"):
        print("seat")
        tk(seat)
cv2.destroyAllWindows()
cap.stop()